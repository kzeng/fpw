import os
import sys
import datetime
import shutil
from intelhex import IntelHex
import json
import struct

def build(PrjDir, json_cfg, nvr_bin_list, nvr_ver, nvr_dict_list):
    '''
    :param PrjDir: project directory path
    :param json_cfg: the content of postbuild.json config file as dict
    :param nvr_bin_list: the list of nvr binary file name to be used for release_wNvr.bin
    :param nvr_ver: nvr version number in U32
    :param nvr_dict_list: the list of dict for nvr data, to be used for JLINK hex/bin generation
    :return: None
    '''
    has_nvr = json_cfg['nvr']['included']
    has_dsp = json_cfg['dspfw']['included']
    encrypt_included = json_cfg['encrypt_included']
    non_encrypt_included = json_cfg['non_encrypt_included']
    dsp_skip_hdr_imgA = json_cfg['dspfw']['skipDspHdrImgA']
    old_releasePrefix = "ses11_qddMain_22335307_001_000" # format as release_prefix_vx_x_x_wodsp.hex
    releasePrefix = "release_GungnirS_22370580_001_000" # format as release_prefix_vx_x_x_wodsp.hex
    old_jlinkHexPrefix = "qddMain_app_AB"
    jlinkHexPrefix = "forJlink_GungnirS_Img"
    release_folder_format = "GungnirS_FW" # format as GungnirS_FW_x.x.x.x_DSP_xxxx
    # format as $(release_prefix)_vx_x_x_wodsp.hex
    # format as $(release_prefix)_vx_x_x_dspv0_0_0.hex
    # MEMORY MAP
    FLASH_STRTADDR   = 0x08000000 # Flash start address, STM32H563:Non-secure:0x8000000, Secure:0xC000000,  ADuCM430: 0x00000000L
    FLASH_SIZE_  = 0x100000 # FLASH_BANK_SIZE, 430IE/STM32 is 1M, 430/410 is 512K
    Gboot_F0_StartAddr  =   FLASH_STRTADDR + 0x00000
    Gboot_F0_EndAddr    =   FLASH_STRTADDR + 0x02000
    CrcTab_F0_StartAddr =   FLASH_STRTADDR + 0x00400  # CrcTab_A
    CrcTab_F0_EndAddr   =   FLASH_STRTADDR + 0x00a00
    Gboot_F1_StartAddr  =   FLASH_SIZE_ + Gboot_F0_StartAddr
    Gboot_F1_EndAddr    =   FLASH_SIZE_ + Gboot_F0_EndAddr
    CrcTab_F1_StartAddr =   FLASH_SIZE_ + CrcTab_F0_StartAddr # CrcTab_B
    CrcTab_F1_EndAddr   =   FLASH_SIZE_ + CrcTab_F0_EndAddr
    VerStr_Gboot_StartAddr  =   FLASH_STRTADDR + 0x00250
    VerStr_IMGA_StartAddr   =   FLASH_STRTADDR + 0x10250
    VerStr_IMGB_StartAddr   =   FLASH_SIZE_ + VerStr_IMGA_StartAddr
    BOOT_CTRL_BLK_IMGA    = (FLASH_STRTADDR + (FLASH_SIZE_ + 0x0000C000))
    BOOT_CTRL_BLK_IMGB    = (FLASH_STRTADDR + 0x0000C000)
    DSPFW_MCUFLASH_STRTADDR_P1 = (FLASH_STRTADDR + 0x00080000)  #DSP internal MCUFLASH IMGA part1 start address
    DSPFW_MCUFLASH_LEN_P1      = 0x80000  #DSP internal MCUFLASH IMGA length part1, 512k
    DSPFW_MCUFLASH_STRTADDR_P2 = (DSPFW_MCUFLASH_STRTADDR_P1 + FLASH_SIZE_)  #DSP internal MCUFLASH IMGA part2 start address
    DSPFW_MCUFLASH_LEN_P2      = 0x13000  #DSP internal MCUFLASH IMGA length part2, 12k
    DSPFW_MCUFLASH_STRTADDR_A  = DSPFW_MCUFLASH_STRTADDR_P1  #DSP internal MCUFLASH IMGA start address
    DSPFW_MCUFLASH_LEN_A       = (DSPFW_MCUFLASH_LEN_P1 + DSPFW_MCUFLASH_LEN_P2)  #DSP internal MCUFLASH IMGA length, 512k(previous), 524k(since fw D0011020).

    # PrjDir = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
    # PrjDir = os.path.normpath(r'C:\Users\xianqing.shao62\ws_Gungnir_S\1p6_sian3_fw_h5\v0_1\SES')
    ToolDir = os.path.abspath(os.path.join(os.getcwd(), "../../tools"))
    qdd2ImgArPath = os.path.join(ToolDir, "imgAr\Release\imgAr.exe")
    # print("PrjDir={}\nToolDir={}\nqdd2ImgArPath={}".format(PrjDir, ToolDir, qdd2ImgArPath))


    # remove old release file
    for file in os.listdir(PrjDir):
        if old_jlinkHexPrefix in file and (".hex" in file or ".bin" in file): # remove old burn file
            os.remove(file)
        if jlinkHexPrefix in file and (".hex" in file or ".bin" in file): # remove old burn file
            os.remove(file)
        if file == "imgA.hex" or file == "imgB.hex" or file == "imgBoot.hex": # remove old copy temp file
            os.remove(file)
        if file.startswith(old_releasePrefix) and file.endswith('.bin'): # remove old release file
            os.remove(file)
        if file.startswith(releasePrefix) and file.endswith('.bin'): # remove old release file
            os.remove(file)
        if (file.startswith("dsp_v") or file.startswith("dspv")) and file.endswith('.bin'): # remove old dsp binary
            os.remove(file)

    shutil.copy(os.path.join(PrjDir, "Output/GungnirS_app/gboot/Exe/GungnirS_app.hex"), "imgBoot.hex")
    shutil.copy(os.path.join(PrjDir, "Output/GungnirS_app/ImageA/Exe/GungnirS_app.hex"), "imgA.hex")
    shutil.copy(os.path.join(PrjDir, "Output/GungnirS_app/ImageB/Exe/GungnirS_app.hex"), "imgB.hex")


    # print("Creating full flash HEX image")
    imgBoot = IntelHex("imgBoot.hex")
    bootStrtAddr = imgBoot.start_addr # store gboot start addr
    imgA = IntelHex("imgA.hex")
    imgB = IntelHex("imgB.hex")

    # skip NVR/CFG DATA
    imgBootNew = imgBoot[Gboot_F0_StartAddr:Gboot_F0_EndAddr]   # imgBoot_part1, 0 ~ 0x02000
    # imgBootNew.merge(imgBoot[CrcTab_F0_StartAddr:CrcTab_F0_EndAddr])  # imgBoot_crcTabA, no need to merge for this mapping(inside gboot)
    imgBootNew.merge(imgBoot[Gboot_F1_StartAddr:Gboot_F1_EndAddr])  # imgBoot_part2, 0x80000~ 0x82000
    # imgBootNew.merge(imgBoot[CrcTab_F1_StartAddr:CrcTab_F1_EndAddr])  # imgBoot_crcTabB, no need to merge for this mapping(inside gboot)

    imgBootNew.merge(imgA,overlap='replace')
    imgBootNew.merge(imgB,overlap='replace')
    # set default committed info for imgA
    imgBootNew.puts(BOOT_CTRL_BLK_IMGA, b'\x00\x00') # set default committed info for imgA
    imgBootNew.start_addr = bootStrtAddr
    gboot_ver = ""
    for byte in imgBootNew[VerStr_Gboot_StartAddr : VerStr_Gboot_StartAddr+0x7].tobinarray():
        gboot_ver += chr(byte)

    imgA_ver = ""
    for byte in imgBootNew[VerStr_IMGA_StartAddr : VerStr_IMGA_StartAddr+0x7].tobinarray():
        imgA_ver += chr(byte)

    imgB_ver = ""
    for byte in imgBootNew[VerStr_IMGB_StartAddr : VerStr_IMGB_StartAddr +0x7].tobinarray():
        imgB_ver += chr(byte)

    print("gboot ver:{}, imgA ver:{}, imgB ver:{}".format(gboot_ver,imgA_ver,imgB_ver))
    if imgA_ver != imgB_ver:
        raise ValueError("WARING: imgA_ver({}) inconsistent with imgB_ver({})")
    # Generate jlink hex/bin file woDSP_woNVR data
    verInfo =imgA_ver.replace('.','_')
    burnHex = '{}_v{}.hex'.format(jlinkHexPrefix, verInfo)
    burnBin = '{}_v{}.bin'.format(jlinkHexPrefix, verInfo)
    burnDspHex = '{}_wdsp_v{}.hex'.format(jlinkHexPrefix, verInfo)
    burnDspBin = '{}_wdsp_v{}.bin'.format(jlinkHexPrefix, verInfo)
    imgBootNew.write_hex_file(open(burnHex,'w'),write_start_addr=True)
    imgBootNew.padding = 0xff
    imgBootNew.tobinfile(open(burnBin,'wb'))
    print("succeed to merge burn hexFile:{}, binFile:{}".format(burnHex, burnBin))

    # search for dsp image file
    dspImgDir = os.path.abspath(os.path.join(os.getcwd(), "../fw/src/dsp/images"))
    dspSrcFile_ImgA = None
    dspBinFile_ImgA = ''
    dspVerInfo = ''
    dspVerU32 = 0
    for file in os.listdir(dspImgDir):
        if "dsp_v" in file:
            dspSrcFile_ImgA = os.path.join(dspImgDir,file)
            suffix = '_ig1_A.bin' if dsp_skip_hdr_imgA else '_ig0_A.bin'
            dspBinFile_ImgA = file.replace('.bin', suffix)
            dspVerInfo = file[5:-4] # "dsp_vE000F200.bin" -> E000F200
            dspVerU32 = int(dspVerInfo, 16)
            break
    if dspSrcFile_ImgA is None:
        print("DSP FILE NOT FOUND")
    else:
        shutil.copy(dspSrcFile_ImgA, dspBinFile_ImgA)

    # Generate jlink hex/bin file wDSP_woNVR data
    if dspSrcFile_ImgA is not None:
        dspImgData = open(dspBinFile_ImgA,'rb').read()
        if len(dspImgData) > DSPFW_MCUFLASH_LEN_A:
            raise ValueError("WARING: dsp image size({}) exceed {}K".format(len(dspImgData), DSPFW_MCUFLASH_LEN_A//1024))
        imgBootNew.puts(DSPFW_MCUFLASH_STRTADDR_P1, dspImgData[:DSPFW_MCUFLASH_LEN_P1]) # write dsp image part1 to imgBootNew
        imgBootNew.puts(DSPFW_MCUFLASH_STRTADDR_P2, dspImgData[DSPFW_MCUFLASH_LEN_P1:DSPFW_MCUFLASH_LEN_P1+DSPFW_MCUFLASH_LEN_P2]) # write dsp image part2 to imgBootNew
        imgBootNew.write_hex_file(open(burnDspHex,'w'),write_start_addr=True)
        imgBootNew.padding = 0xff
        imgBootNew.tobinfile(open(burnDspBin,'wb'))
        print("succeed to merge burn with dsp hexFile:{}, binFile:{}".format(burnDspHex, burnDspBin))
    
    # Generate jlink hex/bin file wDsp_wAANVR data for PCBA loading.
    burnDspAANVRHex_List = []
    PCBA_VER_Dict = {0:'ALPHA_BETA001',  2:'BETA002'} # 0: Alpha/Beta001(Default), 1: Beta001, 2: Beta002, 3-:reserved
    HOUSING_LIST = {'IHS', 'RHS'} 
    for pcbVer, pcbVerName in PCBA_VER_Dict.items():
        for housingType in HOUSING_LIST:
            imgBootNew_temp = imgBootNew
            for nvr_info_dict in nvr_dict_list:
                nvrdata = nvr_info_dict['nvrdata']
                dataLen = nvr_info_dict['dataLen']
                fStrtAddr = nvr_info_dict['fStrtAddr']
                # nvr_ver = nvr_info_dict['nvr_ver']
                if nvr_info_dict['page'] == 255:  #255 is special case, it contain DSP IMAGE INFO PAGE (PAGE 255/2) and NVR INFO PAGE (PAGE 255/3).
                    # DSP IMAGE INFO PAGE (PAGE 255/2)
                    p255_2_data = bytearray([255]*128)
                    p255_2_data[0] = 0xB8   #DSP ImgA is running
                    p255_2_data[4:8] =  struct.pack('<I',  dspVerU32) 
                    p255_2_data[8:12] = struct.pack('<I',  dspVerU32) 
                    nvrdata = changePageBankRegVal(nvrdata, fStrtAddr, dataLen, 255, 2, 128, p255_2_data) # update dsp image info page data according to current DSP version
                    # NVR INFO PAGE (PAGE 255/3)    
                    p255_3_data = bytearray([255]*128)
                    p255_3_data[4:8] =  struct.pack('<I', nvr_ver)
                    p255_3_data[8:12] = struct.pack('<I', nvr_ver)
                    nvrdata = changePageBankRegVal(nvrdata, fStrtAddr, dataLen, 255, 3, 128, p255_3_data) # update nvr info page data according to current nvr version
                if nvr_info_dict['page'] == 254:  # modify several setting for AA
                    # type_mask: bit mask to select which type of NVR sheets to be downloaded
                    # bit0: 0: RHS, 1: IHS
                    # bit1: 0: uncooled, 1: cooled
                    # bit2: 0: withDsp, 1: w/o Dsp
                    # bit3: 0: normal, 1: AA
                    # bit4-: reserved
                    type_mask = 0
                    if housingType == 'IHS':
                        type_mask |= 0x1
                    type_mask |= (1<<3)  # AA type NVR
                    nvrdata = updateP254NvrDataPerType(nvrdata, fStrtAddr, dataLen, pcbVer, type_mask)
                imgBootNew_temp.puts(fStrtAddr, bytes(nvrdata)) # write nvr data to imgBootNew-imgA
                imgBootNew_temp.puts(fStrtAddr + FLASH_SIZE_, bytes(nvrdata)) # write nvr data to imgBootNew-imgB
            verInfo =imgA_ver.replace('.','_')
            burnDspAANVRHex = '{}_wdsp_v{}_wAANVR_{}_{}.hex'.format(jlinkHexPrefix, verInfo, housingType, pcbVerName)
            burnDspAANVRBin = burnDspAANVRHex.replace('.hex','.bin')
            imgBootNew_temp.write_hex_file(open(burnDspAANVRHex,'w'),write_start_addr=True)
            # enable if need analyze the merged data file
            # imgBootNew_temp.padding = 0xff
            # imgBootNew_temp.tobinfile(open(burnDspAANVRBin,'wb'))
            burnDspAANVRHex_List.append(burnDspAANVRHex)
            print("succeed to merge burn hexFile:{}, binFile:{}".format(burnDspAANVRHex, burnDspAANVRBin)) 
   
    dateStr = datetime.datetime.now().strftime("%Y%m%d")
    timeStr = datetime.datetime.now().strftime("%H:%M:%S")
    ReleaseDir = os.path.abspath(os.path.join(os.getcwd(), "../../../Release/{}_{}_DSP_{}".format(release_folder_format, imgA_ver, dspVerInfo)))
    PCBA_FreshLoad_wAANVR_Dir = os.path.join(ReleaseDir, "PCBA_FreshLoad_wAANVR")
    if os.path.exists(ReleaseDir):
        shutil.rmtree(ReleaseDir)
    os.makedirs(ReleaseDir)
    os.makedirs(PCBA_FreshLoad_wAANVR_Dir)

    shutil.copy(burnDspHex, ReleaseDir)
    shutil.copy(burnDspBin, ReleaseDir)
    for burnDspAANVRHex in burnDspAANVRHex_List:
        shutil.copy(burnDspAANVRHex, PCBA_FreshLoad_wAANVR_Dir)
        os.remove(burnDspAANVRHex) # avoid too many files in release dir
    # run imgAr to create CMIS upgrade release binary file
    for i in range(2): # non-encrypt + encrypt version
        if i == 0 and not non_encrypt_included:
            continue
        if i == 1 and not encrypt_included:
            continue
        binFilePath_mcu = "{}_v{}_wodsp.bin".format(releasePrefix,verInfo) if i == 0 else "{}_v{}_enc_wodsp.bin".format(releasePrefix,verInfo)
        if has_dsp:
            binFilePath_mcu_dsp = "{}_v{}_dspv{}.bin".format(releasePrefix,verInfo,dspVerInfo) if i == 0 else "{}_v{}_enc_dspv{}.bin".format(releasePrefix,verInfo,dspVerInfo)
        # if has_nvr: # avoid confusion with too many combinations
        #     binFilePath_mcu_nvr = "{}_v{}_nvrv{:08X}.bin".format(releasePrefix,verInfo,nvr_ver) if i == 0 else "{}_v{}_enc_nvrv{:08X}.bin".format(releasePrefix,verInfo,nvr_ver)
        if has_nvr and has_dsp:
            binFilePath_mcu_dsp_nvr = "{}_v{}_dspv{}_nvrv{:08X}.bin".format(releasePrefix,verInfo,dspVerInfo,nvr_ver) if i == 0 else "{}_v{}_enc_dspv{}_nvrv{:08X}.bin".format(releasePrefix,verInfo,dspVerInfo,nvr_ver)
        encryptType = "enc0" if i==0 else "enc1"
        succeed = 1
        cmd = "{} {} {} {} {} {} {}".format(qdd2ImgArPath, binFilePath_mcu, encryptType, "IMG-A", dateStr,timeStr, "imgA.hex")
        os.system(cmd)
        cmd = "{} {} {} {} {} {} {}".format(qdd2ImgArPath, binFilePath_mcu, encryptType, "IMG-B", dateStr,timeStr, "imgB.hex")
        os.system(cmd)

        try:
            shutil.copy(binFilePath_mcu, binFilePath_mcu_dsp)
        except Exception as e:
            # Ignore the error and continue
            succeed = 0
            print("qdd2ImgAr.exe run error")
            print("failed to release {}".format(binFilePath_mcu))
            print("failed to release {}".format(binFilePath_mcu_dsp))
            pass

        if succeed:
            if has_dsp:
                cmd = "{} {} {} {} {} {} {}".format(qdd2ImgArPath, binFilePath_mcu_dsp, encryptType, "DSP-N-A", dateStr,timeStr, dspBinFile_ImgA)
                os.system(cmd)
                # cmd = "{} {} {} {} {} {} {}".format(qdd2ImgArPath, binFilePath_mcu_dsp, encryptType, "DSP-N-B", dateStr,timeStr, dspBinFile_ImgB)
                # os.system(cmd)
                if has_nvr and has_dsp:
                    shutil.copy(binFilePath_mcu_dsp, binFilePath_mcu_dsp_nvr)
                    for nvr_bin in nvr_bin_list:
                        cmd = "{} {} {} {} {} {} {}".format(qdd2ImgArPath, binFilePath_mcu_dsp_nvr, encryptType, "NVR-REG", dateStr,timeStr, nvr_bin)
                        os.system(cmd)
            shutil.copy(binFilePath_mcu, ReleaseDir)
            os.remove(binFilePath_mcu)
            print("succeed to release {}".format(os.path.join(ReleaseDir, binFilePath_mcu)))
            if has_dsp:
                shutil.copy(binFilePath_mcu_dsp, ReleaseDir)
                os.remove(binFilePath_mcu_dsp)
                print("succeed to release {}".format(os.path.join(ReleaseDir, binFilePath_mcu_dsp)))
            if has_nvr and has_dsp:
                shutil.copy(binFilePath_mcu_dsp_nvr, ReleaseDir)
                os.remove(binFilePath_mcu_dsp_nvr)
                print("succeed to release {}".format(os.path.join(ReleaseDir, binFilePath_mcu_dsp_nvr)))
      
    # remove temp files
    os.remove("imgA.hex")
    os.remove("imgB.hex")
    os.remove("imgBoot.hex")
    os.remove(dspBinFile_ImgA)
    if has_nvr:
        for nvr_bin in nvr_bin_list:
            os.remove(nvr_bin)
    # os.remove(dspBinFile_ImgB)
    # input("Press Enter to continue...")
    #
    # # Running imgAr to create binary file
    # print("Running imgAr to create binary file")

def changePageBankRegVal(nvrDataInOut, nvrDataFlshAddr, nvrDataLen, page, bank, regStrt, regVals):
    if page == 254:
        BaseAddr = 0x8002000
    elif page == 248:
        BaseAddr = 0x8004000
    elif page == 250:
        BaseAddr = 0x8008000
    elif page == 252:
        BaseAddr = 0x8078000
    elif page == 255:
        BaseAddr = 0x807E000
    else:
        raise ValueError("invalid page number: {}".format(page))
    reg_flsh_addr = BaseAddr + bank*128 + regStrt - 128
    if reg_flsh_addr < nvrDataFlshAddr or reg_flsh_addr + len(regVals) > nvrDataFlshAddr + nvrDataLen:
        print("ignore reg addr out of range, reg_flsh_addr: {:#x}, nvrDataFlshAddr: {:#x}, nvrDataLen: {:#x}".format(reg_flsh_addr, nvrDataFlshAddr, nvrDataLen))
        return nvrDataInOut
    for idx, regVal in enumerate(regVals):
        if regVal is not None:
            if not isinstance(regVal, int) or not (0 <= regVal <= 255):
                raise ValueError("regVal {} format error".format(regVal))
            nvrDataInOut[reg_flsh_addr + idx - nvrDataFlshAddr] = regVal
    return nvrDataInOut

def updateP254NvrDataPerType(nvrDataInput, nvrDataFlshAddr, nvrDataLen, pcbVer, type_mask):
    '''
    update p254_x NVR data according to different type of PCBA and setting
    :param nvrDataInput: original NVR data as bytearray
    :param nvrDataFlshAddr: flash address for the NVR data
    :param nvrDataLen: length of the NVR data
    :param pcbVer: PCB version, 0: Alpha/Beta001(Default), 1: Beta001, 2: Beta002, 3-:reserved
    :param type_mask: bit mask to select which type of NVR sheets to be downloaded
                        bit0: 0: RHS, 1: IHS
                        bit1: 0: uncooled, 1: cooled
                        bit2: 0: withDsp, 1: w/o Dsp
                        bit3: 0: normal, 1: AA
                        bit4-: reserved
    :return: modified NVR data as bytearray
    '''
    nvrData = bytearray(nvrDataInput) # create a copy of input data to modify
     #step1: decode type_mask
    isIHS = type_mask & 0x1
    isCooled = type_mask >> 1 & 0x1
    isWoDsp = (type_mask >> 2) & 0x1
    isAA = (type_mask >>3) & 0x1
    if pcbVer in [0,1]: #optional, the same with default NVR setting
        nvrData = changePageBankRegVal(nvrData, nvrDataFlshAddr, nvrDataLen, 254, 0, 134, [0x02]) #enable bit[1]:useVeeAsDspVccCtrl
    elif pcbVer ==2:
        nvrData = changePageBankRegVal(nvrData, nvrDataFlshAddr, nvrDataLen, 254, 0, 134, [0x12]) #enable bit[1]:useVeeAsDspVccCtrl
    if isIHS:  #caseTemp calib coeffs for IHS,       
        nvrData = changePageBankRegVal(nvrData, nvrDataFlshAddr, nvrDataLen, 254,3,128,[61,252,167,61,0,0])
        nvrData = changePageBankRegVal(nvrData, nvrDataFlshAddr, nvrDataLen, 254,9,148,[57,54,53,51,45,69,49,45,49,48,49,49])  #'9653-E1-1011'
    else:
        nvrData = changePageBankRegVal(nvrData, nvrDataFlshAddr, nvrDataLen, 254,3,128,[232,252,70,61,0,0])
        nvrData = changePageBankRegVal(nvrData, nvrDataFlshAddr, nvrDataLen, 254,9,148,[57,54,53,51,45,69,49,45,49,48,49,50])  #'9653-E1-1012'

    if isCooled:   #enable TEC loop
        nvrData = changePageBankRegVal(nvrData, nvrDataFlshAddr, nvrDataLen, 254,1,144,[0x77])
    else:           #disable TEC loop
        nvrData = changePageBankRegVal(nvrData, nvrDataFlshAddr, nvrDataLen, 254,1,144,[0x0])
    
    if isWoDsp:    #disable DSP related settings
        nvrData = changePageBankRegVal(nvrData, nvrDataFlshAddr, nvrDataLen, 254,1,228,[0<<2])
    else:           #enable DSP related settings
        nvrData = changePageBankRegVal(nvrData, nvrDataFlshAddr, nvrDataLen, 254,1,228,[1<<2])
    if isAA:       #AA setting
        nvrData = changePageBankRegVal(nvrData, nvrDataFlshAddr, nvrDataLen, 254, 1, 144, [0]) # AA need to disable TEC loop
        nvrData = changePageBankRegVal(nvrData, nvrDataFlshAddr, nvrDataLen, 254, 1, 147, [0, 0]) # Dis temperature compensation
        nvrData = changePageBankRegVal(nvrData, nvrDataFlshAddr, nvrDataLen, 254, 1, 228, [0]) # Dis TecReadyChk(Dis DSP up)
        nvrData = changePageBankRegVal(nvrData, nvrDataFlshAddr, nvrDataLen, 254, 8, 154, [0x60])  #enLowpwrHW
        nvrData = changePageBankRegVal(nvrData, nvrDataFlshAddr, nvrDataLen, 254, 32, 144, [0]*48)  # All Ibias = 0	
    return nvrData

def loadNvrData(wb, fAddr_strtPB, totLen, sheetList, reg_strt=128, reg_end=255):
    '''
    :param wb: loaded workbook object of openpyxl
    :param fAddr_strtPB: flash address for 'page'-'bank_strt' address 
    :param totLen: total length of NVR data, contain padding, should be n*256 bytes
    :param sheetList: sheet list for specified page/bank range
    :param reg_strt: the start register address for first sheet in sheetList
    :param reg_end: the end register address for last sheet in sheetList
    :return: dict{'nvrdata': bytearray, 'version': int, 'dataLen': int}, dataLen is the length of real NVR data without padding
    '''
    # from openpyxl import load_workbook
    # wb = load_workbook(inputFilename, data_only=True)
    coverSheet = wb['Cover']
    ret = {'nvrdata': None, 'version': None}
    # read version info from cover sheet
    dataverStr = coverSheet['E3'].value
    data_versionNum = struct.unpack(">I", bytes.fromhex(dataverStr))[0]
    print("data versionNum read from cover sheet: {:#x}".format(data_versionNum))
    ret["version"] = data_versionNum
    # #step1: init original data as 0xFF
    nvrData = bytearray('\xFF' * totLen, encoding='latin1')
    #step2: load xlsx data to original data
    fAddr_regStrt = 0
    fAddr_regEnd = 0
    for sheetIdx,sheetInfo in enumerate(sheetList):
        sheetname = sheetInfo['sheetname']
        rowStrt = sheetInfo['rowStrt']
        rowEnd = sheetInfo['rowEnd']
        dataCol = sheetInfo['dataCol']
        page = sheetInfo['page']
        bank = sheetInfo['bank']
        fStrtAddr = sheetInfo['fStrtAddr']
        flen = sheetInfo['flen']
        if sheetname not in wb.sheetnames:
            print("SHEET {} not included in file, skipped".format(sheetname))
            continue
        sheet = wb[sheetname]
        dataSheet = bytearray()
        for rowIdx in range(rowStrt, rowEnd + 1, 1):
            regAddr = sheet[rowIdx][0].value
            if regAddr is None: # skip merge cell
                continue
            regAddr = regAddr + 128 if regAddr < 128 else regAddr # Low page address need to add 128 to get NVR address
            regVal = sheet[rowIdx][dataCol].value
            if not isinstance(regVal, int) or not (0 <= regVal <= 255):
                raise ValueError("sheet {}: row {}, col:{} data format error".format(sheetname,rowIdx,dataCol))
            dataSheet.append(regVal)
        if len(dataSheet)!=flen:
            raise ValueError("{} len({}!={}) ".format(sheetname, len(dataSheet), flen))
        if fStrtAddr < fAddr_strtPB or fStrtAddr + flen > fAddr_strtPB + totLen:
            raise ValueError("sheet {} fStrtAddr out of range, fAddr_strtPB: {:#x}, totLen: {:#x}".format(sheetname, fAddr_strtPB, totLen))
        if sheetIdx == 0: #first sheet for bank_strt
            fAddr_regStrt = fStrtAddr + reg_strt - 128
            if reg_strt > 128:
                dataSheet[:(reg_strt-128)] = b'\xFF'*(reg_strt-128) # padding 0xFF for the data before reg_strt
        if sheetIdx == len(sheetList)-1: #last sheet for bank_end
            fAddr_regEnd = fStrtAddr + reg_end - 128
            if reg_end < 255:
                dataSheet[(reg_end- 255):] = b'\xFF'*(255 - reg_end) # padding 0xFF for the data after reg_end
        nvrData[fStrtAddr-fAddr_strtPB:fStrtAddr-fAddr_strtPB+flen] = dataSheet
        print("sheet {} data load done".format(sheetname))
    #step3: trunc as from reg_strt, then padding FF to totLen.
    dataLen = fAddr_regEnd - fAddr_regStrt + 1 #real data length from bank-strt-reg_strt to bank-end-reg_end
    nvrData = nvrData[reg_strt-128:]
    if len(nvrData) < totLen:
        nvrData.extend(bytearray([0xFF]*(totLen - len(nvrData))))
    ret["nvrdata"] = nvrData
    ret["dataLen"] = dataLen
    return ret

def validate_json_cfg(json_cfg):
    #rule1 :'alt_base' must only be applied once on one Flash Sector, and only for first upgrade block of the sector, 
    #        unless it will overwrite data programmed by previous block(s) of the same sector, then cause CRC issue
    page_sel_list = json_cfg['nvr']['page_sel']
    flash_sector_dict = {} # key: flash sector number, value: whether alt_base is applied for this sector
    for page_sel in page_sel_list:
        sel = page_sel['sel']
        if sel == 0:
            continue
        page = page_sel['page']
        bank_strt = page_sel['bank_strt']
        alt_base = page_sel.get('alt_base', 0)
        BaseAddr= page_sel['BaseAddr'] # flash address for page(bank 0)
        fAddr_strtPB = BaseAddr + bank_strt*128 # flash address for 'page'-'bank_strt' address
        sector_idx = (fAddr_strtPB - 0x08000000)//0x2000 #FLASH_SECTOR_SIZE
        if sector_idx not in flash_sector_dict:
            flash_sector_dict[sector_idx] = [alt_base] #first upgrade block of this sector, allow alt_base == 0 or 1
        else:
            flash_sector_dict[sector_idx].append(alt_base)
            if alt_base == 1:
                print("ERROR: error in postbuild json - 'page_sel'-page{}, bank_strt{}- 'alt_base'".format(page, bank_strt))
                print("ERROR: 'alt_base' only alllowed be applied for the 'first' upgrade block of the flash sector {}:{:#x}, and only once!".format(sector_idx, fAddr_strtPB))
                raise ValueError("invalid json config: alt_base rule violation")
    return True

def genNvrBin(json_cfg, wb=None):
    '''
    generate nvr binary file according to json config, used for release bin packaging
    :param json_cfg: config dict loaded from json file
    :param wb: loaded workbook object of openpyxl, init if None
    :return: [nvr_bin_list, nvr_ver], nvr_bin_list is the list of generated nvr binary file path, nvr_ver is the version read from xlsx cover sheet
    '''
    if wb is None:
        from openpyxl import load_workbook
        inputFilename = json_cfg['nvr']['path']
        print("load workbook from file {}".format(inputFilename))
        wb = load_workbook(inputFilename, data_only=True)
    validate_json_cfg(json_cfg)
    page_sel_list = json_cfg['nvr']['page_sel']
    nvr_bin_list = []
    # pcbVer_list = []         #  0: Alpha/Beta001(Default), 1: Beta001, 2: Beta002, 3-:reserved
    # nvr_type_mask_list = []  #  #   bit0: 0: RHS, 1: IHS
    #                             #   bit1: 0: uncooled, 1: cooled
    #                             #   bit2: 0: withDsp, 1: w/o Dsp
    #                             #   bit3: 0: normal, 1: AA
    #                             #   bit4-: reserved
    # refresh_bin_cache = json_cfg['nvr']['refresh_bin_cache']
    # bin_cache_path = json_cfg['nvr']['bin_cache_path']
    # TODO: update nvr data according to pcbVer and type_mask
    # TODO: use cache bin to speed up if refresh_bin_cache is false and cache exist
    nvr_ver =0 # nvr version should be unique for each release
    for page_sel in page_sel_list:
        sel = page_sel['sel']
        ignore_flshMaskRule = page_sel['ignore_flshMaskRule']
        alt_base = page_sel['alt_base']
        if sel == 0:
            print("page_sel {} not included, skipped".format(page_sel))
            continue
        page = page_sel['page']
        bank_strt = page_sel['bank_strt']
        bank_end = page_sel['bank_end']
        reg_strt = page_sel['reg_strt']
        reg_end = page_sel['reg_end']
        BaseAddr= page_sel['BaseAddr'] # flash address for page(bank 0)
        bank_num = bank_end - bank_strt + 1
        bank_max_limit = page_sel.get('bank_max_limit', None)
        if bank_max_limit is not None and (bank_strt < 0 or bank_end > bank_max_limit):
            raise ValueError("bank_strt {} and bank_end {} exceed max limit {}".format(bank_strt, bank_end, bank_max_limit))
        if reg_strt not in range(128, 256) :
            raise ValueError("reg_strt {} should be in range 128-255".format(reg_strt))
        if reg_end not in range(128, 256) :
            raise ValueError("reg_end {} should be in range 128-255".format(reg_end))
        if reg_strt > reg_end and bank_num == 1:
            raise ValueError("for single bank case, reg_strt {} should be less than or equal with reg_end {}".format(reg_strt, reg_end))
        # if bank_num %2: #release should be padding as multiple of 256 bytes
        #     raise ValueError("bank_strt {} and bank_end {} error, need bank num {} need be multiple of 2".format(bank_strt, bank_end, bank_num)) 
        if bank_num %2: # padding totlen as n*256bytes if bank num is odd
            bank_num += 1
        totLen = bank_num*128 #n*256bytes
        fAddr_strtPB = BaseAddr + bank_strt*128 # flash address for 'page'-'bank_strt' address
        sheetList = []
        #get sheet list for this page/bank range
        for sheet in json_cfg['nvr']['sheet_list']:
            if sheet['page'] == page and sheet['bank'] >= bank_strt and sheet['bank'] <= bank_end:
                sheet['fStrtAddr'] = BaseAddr + sheet['bank']*128
                sheet['flen'] = 128
                sheetList.append(sheet)
        nvrInfo = loadNvrData(wb, fAddr_strtPB, totLen, sheetList, reg_strt, reg_end) # default nvr data for this page/bank range
        nvr_bin = nvrInfo['nvrdata']
        nvr_ver = nvrInfo['version']
        nvr_dataLen = nvrInfo['dataLen']
        alt_base = page_sel.get('alt_base', 0)
        nvr_binary_filename = "nvr_p{:03}_b{:03}_r{:03}_l{:04}_v{:08x}_ig{:d}_alt{:d}.bin".format(page, bank_strt, reg_strt, nvr_dataLen, nvr_ver, ignore_flshMaskRule, alt_base)
        with open(nvr_binary_filename, 'wb') as f_nvr:
            f_nvr.write(nvr_bin)
        print("nvr page {} bank {}-{}, reg {}-{}, data len={}, totlen={} dump to {}".format(page, bank_strt, bank_end, reg_strt, reg_end, nvr_dataLen, totLen, nvr_binary_filename))
        nvr_bin_list.append(nvr_binary_filename)
    return nvr_bin_list, nvr_ver, wb

def gen_nvr_data_for_jlink_hex(json_cfg, wb=None):
    '''
    generate nvr data for merging into jlink hex file according to json config, used for PCBA loading
    :param json_cfg: config dict loaded from json file
    :param wb: loaded workbook object of openpyxl, init if None
    :return: [nvr_dict_list, nvr_ver, wb]
    '''
    if wb is None:
        from openpyxl import load_workbook
        inputFilename = json_cfg['nvr']['path']
        print("load workbook from file {}".format(inputFilename))
        wb = load_workbook(inputFilename, data_only=True)
    page_sel_list = json_cfg['nvr']['page_sel_for_jlink_hex']
    nvr_dict_list = []
    nvr_ver =0
    for page_sel in page_sel_list:
        sel = page_sel['sel']
        if sel == 0:
            print("page_sel {} not included, skipped".format(page_sel))
            continue
        page = page_sel['page']
        bank_strt = page_sel['bank_strt']
        bank_end = page_sel['bank_end']
        bank_num = bank_end - bank_strt + 1
        bank_max_limit = page_sel.get('bank_max_limit', None)
        if bank_max_limit is not None and (bank_strt < 0 or bank_end > bank_max_limit):
            raise ValueError("bank_strt {} and bank_end {} exceed max limit {}".format(bank_strt, bank_end, bank_max_limit))
        fAddr_strtPB = page_sel['BaseAddr'] + bank_strt*128
        dataLen = bank_num*128
        sheetList = []
        #get sheet list for this page/bank range
        for sheet in json_cfg['nvr']['sheet_list']:
            if sheet['page'] == page and sheet['bank'] >= bank_strt and sheet['bank'] <= bank_end:
                sheet['fStrtAddr'] = page_sel['BaseAddr'] + sheet['bank']*128
                sheet['flen'] = 128
                sheetList.append(sheet)
        nvrInfo = loadNvrData(wb, fAddr_strtPB, dataLen, sheetList) # default nvr data for this page/bank range
        nvr_bin = nvrInfo['nvrdata']
        nvr_ver = nvrInfo['version']
        nvr_info_dict  = {"fStrtAddr": fAddr_strtPB, "nvrdata": nvr_bin,"page": page, "bank_strt": bank_strt, "bank_end": bank_end, "dataLen": dataLen, "nvr_ver": nvr_ver, }
        nvr_dict_list.append(nvr_info_dict)
    return nvr_dict_list, nvr_ver, wb

if __name__ == "__main__":
    # f = open("postbuild.json", "r", encoding="utf-8")
    # json_cfg = json.load(f)
    # nvr_bin_list, nvr_ver, wb = genNvrBin(json_cfg)
    # nvr_dict_list, nvr_ver, wb = gen_nvr_data_for_jlink_hex(json_cfg, wb)
    if len(sys.argv) != 2:
        print("Usage: python postbuild.py <PrjDir>")
        sys.exit(1)

    PrjDir = sys.argv[1]
    with open("postbuild.json", "r", encoding="utf-8") as f:
        json_cfg = json.load(f)
        has_nvr = json_cfg['nvr']['included']
        nvr_bin_list =[]
        nvr_ver =0
        nvr_dict_list = []
        if has_nvr:
            nvr_bin_list, nvr_ver, wb = genNvrBin(json_cfg)
        nvr_dict_list, nvr_ver, wb = gen_nvr_data_for_jlink_hex(json_cfg, wb)
        build(PrjDir, json_cfg, nvr_bin_list, nvr_ver, nvr_dict_list)    
    pass